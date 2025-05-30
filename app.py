import os
import tempfile
import openpyxl
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from wtforms import StringField, TextAreaField, DateTimeField, SelectField, SelectMultipleField, BooleanField, SubmitField
from wtforms.validators import DataRequired, Optional
from datetime import datetime
import openpyxl
from icalendar import Calendar, Event
from io import BytesIO
import tempfile

app = Flask(__name__)

# Configurazione dell'applicazione
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or 'pef-esami-secret-key-2023'
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL') or f'sqlite:///{os.path.join(os.path.dirname(os.path.abspath(__file__)), "instance", "esami_pef.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
from flask_migrate import Migrate

migrate = Migrate(app, db)

# Classi di concorso
CLASSI_CONCORSO = [
    ('A001', 'A001 – Arte e immagine nella scuola secondaria di I grado'),
    ('A007', 'A007 – Discipline audiovisive'),
    ('A008', 'A008 – Discipline geometriche, architettura, design d\'arredamento e scenotecnica'),
    ('A011', 'A011 – Discipline letterarie e latino'),
    ('A012', 'A012 – Discipline letterarie negli istituti di istruzione secondaria di II grado'),
    ('A013', 'A013 – Discipline letterarie, latino e greco'),
    ('A017', 'A017 – Disegno e storia dell\'arte'),
    ('A018', 'A018 – Filosofia e scienze umane'),
    ('A019', 'A019 – Filosofia e storia'),
    ('A020', 'A020 – Fisica'),
    ('A022', 'A022 – Italiano, storia, geografia nella scuola secondaria di I grado'),
    ('A023', 'A023 – Lingua italiana per discenti di lingua straniera (alloglotti)'),
    ('A026', 'A026 – Matematica'),
    ('A027', 'A027 – Matematica e fisica'),
    ('A028', 'A028 – Matematica e scienze'),
    ('A029', 'A029 – Musica negli istituti di istruzione secondaria di II grado'),
    ('A030', 'A030 – Musica nella scuola secondaria di I grado'),
    ('A037', 'A037 – Scienze e tecnologie delle costruzioni, tecnologie e tecniche di rappresentazione grafica'),
    ('A040', 'A040 – Tecnologie elettriche elettroniche'),
    ('A042', 'A042 – Scienze e tecnologie meccaniche'),
    ('A045', 'A045 – Scienze economico-aziendali'),
    ('A046', 'A046 – Scienze giuridico-economiche'),
    ('A050', 'A050 – Scienze naturali, chimiche e biologiche'),
    ('A053', 'A053 – Storia della musica'),
    ('A054', 'A054 – Storia dell\'arte'),
    ('A060', 'A060 – Tecnologia nella scuola secondaria di I grado'),
    ('A061', 'A061 – Tecnologie e tecniche delle comunicazioni multimediali'),
    ('A063', 'A063 – Tecnologie musicali'),
    ('A064', 'A064 – Teoria, analisi e composizione'),
    ('AA24', 'AA24 – Lingua e cultura straniera (francese)'),
    ('AB24', 'AB24 – Lingua e cultura straniera (inglese)'),
    ('AC24', 'AC24 – Lingua e cultura straniera (spagnolo)'),
    ('AL24', 'AL24 – Lingua e cultura straniera (arabo)'),
    ('B015', 'B015 – Laboratori di scienze e tecnologie elettriche ed elettroniche')
]

# Percorsi PEF - Rimossi quelli non necessari
PERCORSI_PEF = [
    ('pef_60_all1', 'PeF 60 CFU All. 1 - DPCM 4 ago 2023')
]

# Profili per i membri di commissione
PROFILI_COMMISSIONE = [
    ('prof_ricercatore_roma3', 'Professore / Ricercatore Roma Tre'),
    ('assegnista_roma3', 'Assegnista di ricerca Roma Tre'),
    ('cultore_roma3', 'Cultore della materia Roma Tre'),
    ('docente_contratto', 'Docente a contratto'),
    ('tutor_coordinatore', 'Tutor coordinatore'),
    ('referente_usr', 'Referente USR'),
    ('esterno_altro', 'Esterno (altro)')
]

# Ruoli per i membri di commissione
RUOLI_COMMISSIONE = [
    ('presidente', 'Presidente'),
    ('docente_interno', 'Docente interno'),
    ('esperto_esterno', 'Esperto esterno'),
    ('referente_usr', 'Referente USR')
]

# Tipologie membro commissione
TIPOLOGIE_COMMISSIONE = [
    ('T', 'Titolare'),
    ('S', 'Supplente')
]

# Tipi di attività per le date aggiuntive
TIPI_ATTIVITA = [
    ('definizione_argomento_scritta', 'Definizione argomento prova scritta'),
    ('definizione_argomento_lezione', 'Definizione argomento lezione simulata'),
    ('definizione_criteri', 'Definizione criteri e griglie di valutazione'),
    ('prova_scritta_presenza', 'Prova scritta (in presenza)'),
    ('prova_scritta_consegna', 'Prova scritta (consegna)'),
    ('valutazione_prova_scritta', 'Valutazione prova scritta'),
    ('lezione_simulata', 'Lezione simulata (in presenza)'),
    ('altro', 'Altro')
]

# Modelli del database
class Esame(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    classe_concorso = db.Column(db.String(10), nullable=False)
    percorsi_pef = db.Column(db.Text, nullable=False)  # Cambio da percorso_pef a percorsi_pef per supportare multipli
    modalita = db.Column(db.String(20), nullable=False)  # 'online' o 'presenza'
    sede = db.Column(db.String(200))
    note = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    data_inizio = db.Column(db.DateTime)  # Aggiunto campo data_inizio
    data_fine = db.Column(db.DateTime)    # Aggiunto campo data_fine
    
    # Relazione con il calendario esami (ex date aggiuntive)
    calendario_esami = db.relationship('CalendarioEsame', backref='esame', lazy=True, cascade='all, delete-orphan')
    commissioni = db.relationship('Commissione', backref='esame', lazy=True, cascade='all, delete-orphan')
    
    def get_percorsi_list(self):
        """Restituisce la lista dei percorsi PEF"""
        return self.percorsi_pef.split(',') if self.percorsi_pef else []
    
    def set_percorsi_list(self, percorsi_list):
        """Imposta la lista dei percorsi PEF"""
        self.percorsi_pef = ','.join(percorsi_list) if percorsi_list else ''

class CalendarioEsame(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    esame_id = db.Column(db.Integer, db.ForeignKey('esame.id'), nullable=False)
    data_inizio = db.Column(db.DateTime, nullable=False)
    data_fine = db.Column(db.DateTime, nullable=False)
    modalita = db.Column(db.String(20), nullable=False)
    sede = db.Column(db.String(200))
    attivita = db.Column(db.String(50), nullable=False, default='altro')  # Tipo di attività

class MembroCommissione(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    cognome = db.Column(db.String(100), nullable=False)
    # Il ruolo ora è specifico per ogni commissione
    profilo = db.Column(db.String(100), nullable=False)  # Profili (prof_ricercatore_roma3, assegnista_roma3, ecc.)
    email = db.Column(db.String(120))
    telefono = db.Column(db.String(20))
    note = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relazione con le commissioni
    commissioni = db.relationship('Commissione', backref='membro', lazy=True)
    
    def __repr__(self):
        return f'{self.nome} {self.cognome}'
    
    def calcola_ore_totali(self):
        """Calcola le ore totali di partecipazione alle commissioni"""
        ore_totali = 0
        for commissione in self.commissioni:
            for data in commissione.esame.calendario_esami:
                # Calcola la durata in ore
                durata = (data.data_fine - data.data_inizio).total_seconds() / 3600
                ore_totali += durata
        return round(ore_totali, 1)  # Arrotonda a 1 decimale
    
    def get_esami(self):
        """Restituisce gli esami unici in cui partecipa"""
        esami = []
        esami_ids = []
        for commissione in self.commissioni:
            if commissione.esame_id not in esami_ids:
                esami.append(commissione.esame)
                esami_ids.append(commissione.esame_id)
        return esami

class Commissione(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    esame_id = db.Column(db.Integer, db.ForeignKey('esame.id'), nullable=False)
    membro_id = db.Column(db.Integer, db.ForeignKey('membro_commissione.id'), nullable=True)  # Nuovo campo
    nome_membro = db.Column(db.String(100), nullable=False)
    cognome_membro = db.Column(db.String(100), nullable=False)
    tipologia = db.Column(db.String(10), nullable=False, default='T')  # 'T' = Titolare, 'S' = Supplente
    ruolo = db.Column(db.String(50), nullable=False)  # Ruoli aggiornati
    profilo = db.Column(db.String(100), nullable=False)  # Profili aggiornati
    email = db.Column(db.String(120))
    telefono = db.Column(db.String(20))
    note = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Form per la gestione degli esami
class EsameForm(FlaskForm):
    classe_concorso = SelectField('Classe di Concorso', choices=CLASSI_CONCORSO, validators=[DataRequired()])
    percorsi_pef = SelectMultipleField('Percorsi PEF', choices=PERCORSI_PEF, validators=[DataRequired()])
    data_inizio = DateTimeField('Data e Ora Inizio', validators=[Optional()], format='%Y-%m-%dT%H:%M')
    data_fine = DateTimeField('Data e Ora Fine', validators=[Optional()], format='%Y-%m-%dT%H:%M')
    modalita = SelectField('Modalità', choices=[('online', 'Online'), ('presenza', 'In Presenza'), ('mista', 'Mista')], validators=[DataRequired()])
    sede = StringField('Sede', validators=[Optional()])
    note = TextAreaField('Note', validators=[Optional()])
    submit = SubmitField('Salva Esame')

class CalendarioEsameForm(FlaskForm):
    data_inizio = DateTimeField('Data e Ora Inizio', validators=[DataRequired()], format='%Y-%m-%dT%H:%M')
    data_fine = DateTimeField('Data e Ora Fine', validators=[DataRequired()], format='%Y-%m-%dT%H:%M')
    modalita = SelectField('Modalità', choices=[('online', 'Online'), ('presenza', 'In Presenza'), ('mista', 'Mista')], validators=[DataRequired()])
    sede = StringField('Sede', validators=[Optional()])
    attivita = SelectField('Tipo di Attività', choices=TIPI_ATTIVITA, validators=[DataRequired()])

class CommissioneForm(FlaskForm):
    nome_membro = StringField('Nome', validators=[DataRequired()])
    cognome_membro = StringField('Cognome', validators=[DataRequired()])
    tipologia = SelectField('Tipologia', choices=TIPOLOGIE_COMMISSIONE, validators=[DataRequired()])
    ruolo = SelectField('Ruolo', choices=RUOLI_COMMISSIONE, validators=[DataRequired()])
    profilo = SelectField('Profilo', choices=PROFILI_COMMISSIONE, validators=[DataRequired()])
    email = StringField('Email', validators=[Optional()])
    telefono = StringField('Telefono', validators=[Optional()])
    note = TextAreaField('Note', validators=[Optional()])

class MembroCommissioneForm(FlaskForm):
    """Form per la gestione dei membri delle commissioni"""
    nome = StringField('Nome', validators=[DataRequired()])
    cognome = StringField('Cognome', validators=[DataRequired()])
    profilo = SelectField('Profilo', choices=PROFILI_COMMISSIONE, validators=[DataRequired()])
    email = StringField('Email', validators=[Optional()])
    telefono = StringField('Telefono', validators=[Optional()])
    note = TextAreaField('Note', validators=[Optional()])
    submit = SubmitField('Salva Membro Commissione')

# Routes
@app.route('/')
def index():
    # Parametri di ricerca e filtro
    search = request.args.get('search', '')
    classe_filter = request.args.get('classe', '')
    modalita_filter = request.args.get('modalita', '')
    
    # Query base
    query = Esame.query
    
    # Applicazione filtri
    if search:
        query = query.filter(
            db.or_(
                Esame.classe_concorso.contains(search),
                Esame.sede.contains(search),
                Esame.note.contains(search)
            )
        )
    
    if classe_filter:
        query = query.filter(Esame.classe_concorso == classe_filter)
    
    if modalita_filter:
        query = query.filter(Esame.modalita == modalita_filter)
    
    # Usa created_at per l'ordinamento per essere sicuri
    try:
        esami = query.order_by(Esame.created_at.desc()).all()
    except Exception as e:
        print(f"Errore nell'ordinamento: {e}")
        # Fallback: carica senza ordinamento
        esami = query.all()
    
    # Statistiche per il dashboard
    stats = {
        'totale_esami': Esame.query.count(),
        'online': Esame.query.filter_by(modalita='online').count(),
        'presenza': Esame.query.filter_by(modalita='presenza').count(),
        'mista': Esame.query.filter_by(modalita='mista').count(),
        'con_commissione': Esame.query.join(Commissione).distinct().count(),
        # Statistiche per i ruoli delle commissioni
        'commissioni': {
            'titolari': Commissione.query.filter_by(tipologia='T').count(),
            'supplenti': Commissione.query.filter_by(tipologia='S').count(),
            'presidenti': Commissione.query.filter_by(ruolo='presidente').count(),
            'docenti_interni': Commissione.query.filter_by(ruolo='docente_interno').count(),
            'esperti_esterni': Commissione.query.filter_by(ruolo='esperto_esterno').count(),
            'referenti_usr': Commissione.query.filter_by(ruolo='referente_usr').count()
        },
        # Statistiche per i profili
        'profili': {
            'prof_ricercatori': Commissione.query.filter_by(profilo='prof_ricercatore_roma3').count(),
            'assegnisti': Commissione.query.filter_by(profilo='assegnista_roma3').count(),
            'cultori': Commissione.query.filter_by(profilo='cultore_roma3').count(),
            'docenti_contratto': Commissione.query.filter_by(profilo='docente_contratto').count(),
            'tutor': Commissione.query.filter_by(profilo='tutor_coordinatore').count(),
            'ref_usr': Commissione.query.filter_by(profilo='referente_usr').count(),
            'esterni': Commissione.query.filter_by(profilo='esterno_altro').count()
        },
        # Statistiche per i tipi di attività
        'tipi_attivita': {
            'definizione_argomento_scritta': CalendarioEsame.query.filter_by(attivita='definizione_argomento_scritta').count(),
            'definizione_argomento_lezione': CalendarioEsame.query.filter_by(attivita='definizione_argomento_lezione').count(),
            'definizione_criteri': CalendarioEsame.query.filter_by(attivita='definizione_criteri').count(),
            'prova_scritta_presenza': CalendarioEsame.query.filter_by(attivita='prova_scritta_presenza').count(),
            'prova_scritta_consegna': CalendarioEsame.query.filter_by(attivita='prova_scritta_consegna').count(),
            'valutazione_prova_scritta': CalendarioEsame.query.filter_by(attivita='valutazione_prova_scritta').count(),
            'lezione_simulata': CalendarioEsame.query.filter_by(attivita='lezione_simulata').count(),
            'altro': CalendarioEsame.query.filter_by(attivita='altro').count()
        }
    }
    
    return render_template('index.html', 
                         esami=esami, 
                         stats=stats,
                         classi_concorso=CLASSI_CONCORSO,
                         search=search,
                         classe_filter=classe_filter,
                         modalita_filter=modalita_filter)

@app.route('/nuovo_esame', methods=['GET', 'POST'])
def nuovo_esame():
    form = EsameForm()
    if form.validate_on_submit():
        # Gestione percorsi multipli dal form WTF
        percorsi_selezionati = form.percorsi_pef.data
        percorsi_string = ','.join(percorsi_selezionati) if percorsi_selezionati else ''
        
        esame = Esame(
            classe_concorso=form.classe_concorso.data,
            percorsi_pef=percorsi_string,
            data_inizio=form.data_inizio.data,
            data_fine=form.data_fine.data,
            modalita=form.modalita.data,
            sede=form.sede.data,
            note=form.note.data
        )
        db.session.add(esame)
        db.session.flush()  # Per ottenere l'ID dell'esame
        
        # Gestione date aggiuntive con campi separati
        contatore = 1
        while True:
            data_key = f'data_aggiuntiva_data_{contatore}'
            ora_inizio_key = f'data_aggiuntiva_ora_inizio_{contatore}'
            ora_fine_key = f'data_aggiuntiva_ora_fine_{contatore}'
            modalita_key = f'data_aggiuntiva_modalita_{contatore}'
            sede_key = f'data_aggiuntiva_sede_{contatore}'
            attivita_key = f'data_aggiuntiva_attivita_{contatore}'
            
            if data_key not in request.form:
                break
                
            try:
                data = request.form[data_key]
                ora_inizio = request.form[ora_inizio_key]
                ora_fine = request.form[ora_fine_key]
                modalita = request.form[modalita_key]
                sede = request.form.get(sede_key, '')
                attivita = request.form.get(attivita_key, 'altro')
                
                # Combina data e orari per creare datetime
                data_inizio = datetime.fromisoformat(f"{data}T{ora_inizio}")
                data_fine = datetime.fromisoformat(f"{data}T{ora_fine}")
                
                data_aggiuntiva = CalendarioEsame(
                    esame_id=esame.id,
                    data_inizio=data_inizio,
                    data_fine=data_fine,
                    modalita=modalita,
                    sede=sede,
                    attivita=attivita
                )
                db.session.add(data_aggiuntiva)
            except (ValueError, KeyError):
                pass  # Salta date non valide
            
            contatore += 1

        # Gestione membri commissione
        contatore_commissione = 1
        while True:
            nome_key = f'commissione_nome_{contatore_commissione}'
            cognome_key = f'commissione_cognome_{contatore_commissione}'
            ruolo_key = f'commissione_ruolo_{contatore_commissione}'
            
            if nome_key not in request.form:
                break
                
            try:
                nome = request.form[nome_key].strip()
                cognome = request.form[cognome_key].strip()
                ruolo = request.form[ruolo_key]
                
                if nome and cognome and ruolo:  # Solo se i campi obbligatori sono presenti
                    tipologia_key = f'commissione_tipologia_{contatore_commissione}'
                    profilo_key = f'commissione_profilo_{contatore_commissione}'
                    email_key = f'commissione_email_{contatore_commissione}'
                    telefono_key = f'commissione_telefono_{contatore_commissione}'
                    note_key = f'commissione_note_{contatore_commissione}'
                    
                    tipologia = request.form.get(tipologia_key, 'T')
                    profilo = request.form.get(profilo_key, 'esterno_altro')
                    email = request.form.get(email_key, '')
                    telefono = request.form.get(telefono_key, '')
                    note = request.form.get(note_key, '')
                    
                    membro_commissione = Commissione(
                        esame_id=esame.id,
                        nome_membro=nome,
                        cognome_membro=cognome,
                        tipologia=tipologia,
                        ruolo=ruolo,
                        profilo=profilo,
                        email=email,
                        telefono=telefono,
                        note=note
                    )
                    db.session.add(membro_commissione)
            except (ValueError, KeyError):
                pass  # Salta membri non validi
            
            contatore_commissione += 1
        
        db.session.commit()
        flash('Esame creato con successo!', 'success')
        return redirect(url_for('dettaglio_esame', id=esame.id))
    
    # In caso di errori di validazione, gestisci anche i percorsi PEF da request.form se il form WTF fallisce
    if request.method == 'POST' and not form.validate_on_submit():
        # Gestione alternativa per i percorsi PEF tramite checkbox
        percorsi_selezionati = []
        for percorso_id, _ in PERCORSI_PEF:
            if request.form.get(f'percorso_{percorso_id}'):
                percorsi_selezionati.append(percorso_id)
        
        if percorsi_selezionati:
            percorsi_string = ','.join(percorsi_selezionati)
            
            try:
                esame = Esame(
                    classe_concorso=request.form['classe_concorso'],
                    percorsi_pef=percorsi_string,
                    data_inizio=datetime.fromisoformat(request.form['data_inizio']),
                    data_fine=datetime.fromisoformat(request.form['data_fine']),
                    modalita=request.form['modalita'],
                    sede=request.form.get('sede', ''),
                    note=request.form.get('note', '')
                )
                db.session.add(esame)
                db.session.flush()
                
                # Gestione date aggiuntive
                contatore = 1
                while True:
                    data_key = f'data_aggiuntiva_data_{contatore}'
                    ora_inizio_key = f'data_aggiuntiva_ora_inizio_{contatore}'
                    ora_fine_key = f'data_aggiuntiva_ora_fine_{contatore}'
                    modalita_key = f'data_aggiuntiva_modalita_{contatore}'
                    sede_key = f'data_aggiuntiva_sede_{contatore}'
                    attivita_key = f'data_aggiuntiva_attivita_{contatore}'
                    
                    if data_key not in request.form:
                        break
                        
                    try:
                        data = request.form[data_key]
                        ora_inizio = request.form[ora_inizio_key]
                        ora_fine = request.form[ora_fine_key]
                        modalita = request.form[modalita_key]
                        sede = request.form.get(sede_key, '')
                        attivita = request.form.get(attivita_key, 'altro')
                        
                        data_inizio = datetime.fromisoformat(f"{data}T{ora_inizio}")
                        data_fine = datetime.fromisoformat(f"{data}T{ora_fine}")
                        
                        data_aggiuntiva = CalendarioEsame(
                            esame_id=esame.id,
                            data_inizio=data_inizio,
                            data_fine=data_fine,
                            modalita=modalita,
                            sede=sede,
                            attivita=attivita
                        )
                        db.session.add(data_aggiuntiva)
                    except (ValueError, KeyError):
                        pass
                    
                    contatore += 1

                # Gestione membri commissione
                contatore_commissione = 1
                while True:
                    nome_key = f'commissione_nome_{contatore_commissione}'
                    cognome_key = f'commissione_cognome_{contatore_commissione}'
                    ruolo_key = f'commissione_ruolo_{contatore_commissione}'
                    
                    if nome_key not in request.form:
                        break
                        
                    try:
                        nome = request.form[nome_key].strip()
                        cognome = request.form[cognome_key].strip()
                        ruolo = request.form[ruolo_key]
                        
                        if nome and cognome and ruolo:  # Solo se i campi obbligatori sono presenti
                            profilo_key = f'commissione_profilo_{contatore_commissione}'
                            tipologia_key = f'commissione_tipologia_{contatore_commissione}'
                            email_key = f'commissione_email_{contatore_commissione}'
                            telefono_key = f'commissione_telefono_{contatore_commissione}'
                            note_key = f'commissione_note_{contatore_commissione}'
                            
                            profilo = request.form.get(profilo_key, 'esterno_altro')
                            tipologia = request.form.get(tipologia_key, 'T')
                            email = request.form.get(email_key, '')
                            telefono = request.form.get(telefono_key, '')
                            note = request.form.get(note_key, '')
                            
                            membro_commissione = Commissione(
                                esame_id=esame.id,
                                nome_membro=nome,
                                cognome_membro=cognome,
                                tipologia=tipologia,
                                ruolo=ruolo,
                                profilo=profilo,
                                email=email,
                                telefono=telefono,
                                note=note
                            )
                            db.session.add(membro_commissione)
                    except (ValueError, KeyError):
                        pass  # Salta membri non validi
                    
                    contatore_commissione += 1
                
                db.session.commit()
                flash('Esame creato con successo!', 'success')
                return redirect(url_for('dettaglio_esame', id=esame.id))
                
            except Exception as e:
                db.session.rollback()
                flash(f'Errore durante la creazione dell\'esame: {str(e)}', 'danger')
    
    return render_template('nuovo_esame.html', form=form, 
                         TIPI_ATTIVITA=TIPI_ATTIVITA, 
                         RUOLI_COMMISSIONE=RUOLI_COMMISSIONE,
                         TIPOLOGIE_COMMISSIONE=TIPOLOGIE_COMMISSIONE,
                         PROFILI_COMMISSIONE=PROFILI_COMMISSIONE)

@app.route('/esame/<int:id>')
def dettaglio_esame(id):
    esame = Esame.query.get_or_404(id)
    return render_template('dettaglio_esame.html', esame=esame, 
                         TIPI_ATTIVITA=TIPI_ATTIVITA,
                         RUOLI_COMMISSIONE=RUOLI_COMMISSIONE,
                         TIPOLOGIE_COMMISSIONE=TIPOLOGIE_COMMISSIONE,
                         PROFILI_COMMISSIONE=PROFILI_COMMISSIONE)

@app.route('/esame/<int:id>/modifica', methods=['GET', 'POST'])
def modifica_esame(id):
    esame = Esame.query.get_or_404(id)
    form = EsameForm()
    
    if request.method == 'GET':
        # Precompila il form con i dati esistenti
        form.classe_concorso.data = esame.classe_concorso
        form.percorsi_pef.data = esame.get_percorsi_list()
        form.data_inizio.data = esame.data_inizio
        form.data_fine.data = esame.data_fine
        form.modalita.data = esame.modalita
        form.sede.data = esame.sede
        form.note.data = esame.note
    
    if form.validate_on_submit():
        esame.classe_concorso = form.classe_concorso.data
        esame.set_percorsi_list(form.percorsi_pef.data)
        esame.data_inizio = form.data_inizio.data
        esame.data_fine = form.data_fine.data
        esame.modalita = form.modalita.data
        esame.sede = form.sede.data
        esame.note = form.note.data
        
        # Gestione nuove date aggiuntive (aggiunte tramite il form di modifica)
        contatore = 1
        while True:
            data_key = f'data_aggiuntiva_data_{contatore}'
            ora_inizio_key = f'data_aggiuntiva_ora_inizio_{contatore}'
            ora_fine_key = f'data_aggiuntiva_ora_fine_{contatore}'
            modalita_key = f'data_aggiuntiva_modalita_{contatore}'
            sede_key = f'data_aggiuntiva_sede_{contatore}'
            attivita_key = f'data_aggiuntiva_attivita_{contatore}'
            
            if data_key not in request.form:
                break
                
            try:
                data = request.form[data_key]
                ora_inizio = request.form[ora_inizio_key]
                ora_fine = request.form[ora_fine_key]
                modalita = request.form[modalita_key]
                sede = request.form.get(sede_key, '')
                attivita = request.form.get(attivita_key, 'altro')
                
                # Combina data e orari per creare datetime
                data_inizio = datetime.fromisoformat(f"{data}T{ora_inizio}")
                data_fine = datetime.fromisoformat(f"{data}T{ora_fine}")
                
                data_aggiuntiva = CalendarioEsame(
                    esame_id=esame.id,
                    data_inizio=data_inizio,
                    data_fine=data_fine,
                    modalita=modalita,
                    sede=sede,
                    attivita=attivita
                )
                db.session.add(data_aggiuntiva)
            except (ValueError, KeyError):
                pass  # Salta date non valide
            
            contatore += 1
        
        db.session.commit()
        flash('Esame modificato con successo!', 'success')
        return redirect(url_for('dettaglio_esame', id=esame.id))
    return render_template('modifica_esame.html', form=form, esame=esame, 
                         TIPI_ATTIVITA=TIPI_ATTIVITA,
                         RUOLI_COMMISSIONE=RUOLI_COMMISSIONE,
                         TIPOLOGIE_COMMISSIONE=TIPOLOGIE_COMMISSIONE,
                         PROFILI_COMMISSIONE=PROFILI_COMMISSIONE)

@app.route('/esame/<int:id>/elimina', methods=['POST'])
def elimina_esame(id):
    esame = Esame.query.get_or_404(id)
    db.session.delete(esame)
    db.session.commit()
    flash('Esame eliminato con successo!', 'success')
    return redirect(url_for('index'))

@app.route('/esame/<int:id>/aggiungi_data', methods=['POST'])
def aggiungi_data(id):
    esame = Esame.query.get_or_404(id)
    data_inizio = datetime.fromisoformat(request.form['data_inizio'])
    data_fine = datetime.fromisoformat(request.form['data_fine'])
    modalita = request.form['modalita']
    sede = request.form.get('sede', '')
    attivita = request.form.get('attivita', 'altro')
    
    data_aggiuntiva = CalendarioEsame(
        esame_id=id,
        data_inizio=data_inizio,
        data_fine=data_fine,
        modalita=modalita,
        sede=sede,
        attivita=attivita
    )
    db.session.add(data_aggiuntiva)
    db.session.commit()
    flash('Data aggiuntiva aggiunta!', 'success')
    return redirect(url_for('dettaglio_esame', id=id))

@app.route('/data_aggiuntiva/<int:id>/elimina', methods=['POST'])
def elimina_data_aggiuntiva(id):
    data = CalendarioEsame.query.get_or_404(id)
    esame_id = data.esame_id
    db.session.delete(data)
    db.session.commit()
    flash('Data aggiuntiva eliminata!', 'success')
    return redirect(url_for('dettaglio_esame', id=esame_id))

@app.route('/esame/<int:id>/aggiungi_commissione', methods=['POST'])
def aggiungi_commissione(id):
    esame = Esame.query.get_or_404(id)
    nome_membro = request.form['nome_membro']
    cognome_membro = request.form['cognome_membro']
    tipologia = request.form.get('tipologia', 'T')
    ruolo = request.form['ruolo']
    profilo = request.form['profilo']
    email = request.form.get('email', '')
    telefono = request.form.get('telefono', '')
    note = request.form.get('note', '')
    
    commissione = Commissione(
        esame_id=id,
        nome_membro=nome_membro,
        cognome_membro=cognome_membro,
        tipologia=tipologia,
        ruolo=ruolo,
        profilo=profilo,
        email=email,
        telefono=telefono,
        note=note
    )
    db.session.add(commissione)
    db.session.commit()
    flash('Membro commissione aggiunto!', 'success')
    return redirect(url_for('dettaglio_esame', id=id))

@app.route('/commissione/<int:id>/elimina', methods=['POST'])
def elimina_commissione(id):
    commissione = Commissione.query.get_or_404(id)
    esame_id = commissione.esame_id
    db.session.delete(commissione)
    db.session.commit()
    flash('Membro commissione eliminato!', 'success')
    return redirect(url_for('dettaglio_esame', id=esame_id))

@app.route('/commissione/<int:id>/modifica', methods=['GET', 'POST'])
def modifica_commissione(id):
    commissione = Commissione.query.get_or_404(id)
    form = CommissioneForm(obj=commissione)
    
    if form.validate_on_submit():
        commissione.nome_membro = form.nome_membro.data
        commissione.cognome_membro = form.cognome_membro.data
        commissione.tipologia = form.tipologia.data
        commissione.ruolo = form.ruolo.data
        commissione.profilo = form.profilo.data
        commissione.email = form.email.data
        commissione.telefono = form.telefono.data
        commissione.note = form.note.data
        
        db.session.commit()
        flash('Membro commissione modificato con successo!', 'success')
        return redirect(url_for('dettaglio_esame', id=commissione.esame_id))
    
    return render_template('modifica_commissione.html', form=form, commissione=commissione)

@app.route('/scarica_calendario')
def scarica_calendario():
    cal = Calendar()
    cal.add('prodid', '-//Esami PEF//mxm.dk//')
    cal.add('version', '2.0')
    
    esami = Esame.query.all()
    for esame in esami:
        # Evento principale
        percorsi_nomi = [dict(PERCORSI_PEF)[p] for p in esame.get_percorsi_list()]
        percorsi_str = ', '.join(percorsi_nomi)
        
        # Verifica che esame abbia data_inizio e data_fine
        if hasattr(esame, 'data_inizio') and esame.data_inizio:
            event = Event()
            event.add('summary', f"Esame {esame.classe_concorso} - {percorsi_str}")
            event.add('dtstart', esame.data_inizio)
            if hasattr(esame, 'data_fine') and esame.data_fine:
                event.add('dtend', esame.data_fine)
            if esame.sede:
                event.add('location', esame.sede)
            if esame.note:
                event.add('description', esame.note)
            cal.add_component(event)
        
        # Date aggiuntive (ora chiamate calendario_esami)
        if hasattr(esame, 'calendario_esami'):
            for data_add in esame.calendario_esami:
                event_add = Event()
                event_add.add('summary', f"Esame {esame.classe_concorso} - {percorsi_str} (Data aggiuntiva)")
                event_add.add('dtstart', data_add.data_inizio)
                event_add.add('dtend', data_add.data_fine)
                if data_add.sede:
                    event_add.add('location', data_add.sede)
                if hasattr(data_add, 'attivita') and data_add.attivita:
                    event_add.add('description', dict(TIPI_ATTIVITA).get(data_add.attivita, 'Altro'))
                cal.add_component(event_add)
    
    # Salva il file temporaneo
    with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.ics') as f:
        f.write(cal.to_ical())
        temp_path = f.name
    
    return send_file(temp_path, as_attachment=True, download_name='esami_pef.ics', mimetype='text/calendar')

@app.route('/scarica_excel')
def scarica_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Esami PEF"
    
    # Headers
    headers = ['ID', 'Classe Concorso', 'Percorso PEF', 'Data Inizio', 'Data Fine', 'Modalità', 'Sede', 'Note', 'Commissione', 'Date Aggiuntive']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    esami = Esame.query.all()
    for row, esame in enumerate(esami, 2):
        ws.cell(row=row, column=1, value=esame.id)
        ws.cell(row=row, column=2, value=dict(CLASSI_CONCORSO).get(esame.classe_concorso, esame.classe_concorso))
        
        # Gestione percorsi multipli
        percorsi_nomi = [dict(PERCORSI_PEF).get(p, p) for p in esame.get_percorsi_list()]
        percorsi_str = ', '.join(percorsi_nomi)
        ws.cell(row=row, column=3, value=percorsi_str)

        # Verifica che data_inizio e data_fine esistano
        if hasattr(esame, 'data_inizio') and esame.data_inizio:
            ws.cell(row=row, column=4, value=esame.data_inizio.strftime('%d/%m/%Y %H:%M'))
        else:
            ws.cell(row=row, column=4, value='Non definita')
            
        if hasattr(esame, 'data_fine') and esame.data_fine:
            ws.cell(row=row, column=5, value=esame.data_fine.strftime('%d/%m/%Y %H:%M'))
        else:
            ws.cell(row=row, column=5, value='Non definita')
            
        ws.cell(row=row, column=6, value=esame.modalita)
        ws.cell(row=row, column=7, value=esame.sede or '')
        ws.cell(row=row, column=8, value=esame.note or '')
        
        # Commissione
        commissione_str = '; '.join([
            f"{c.nome_membro} {c.cognome_membro} ({c.ruolo}, {c.tipologia})"
            for c in esame.commissioni
        ])
        ws.cell(row=row, column=9, value=commissione_str)
        
        # Date aggiuntive (ora chiamate calendario_esami)
        calendario = getattr(esame, 'calendario_esami', [])
        date_agg_str = '; '.join([
            f"{d.data_inizio.strftime('%d/%m/%Y %H:%M')} - {d.data_fine.strftime('%d/%m/%Y %H:%M')} ({d.modalita}) - {dict(TIPI_ATTIVITA).get(getattr(d, 'attivita', 'altro'), 'Non specificato')}" 
            for d in calendario
        ])
        ws.cell(row=row, column=10, value=date_agg_str)
    
    # Salva il file temporaneo
    with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xlsx') as f:
        wb.save(f.name)
        temp_path = f.name
    
    return send_file(temp_path, as_attachment=True, download_name='esami_pef.xlsx', 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/scarica_excel_date')
def scarica_excel_date():
    """Genera un file Excel con tutte le date degli esami"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Date Esami PEF"
    
    # Headers
    headers = ['Data', 'Ora Inizio', 'Ora Fine', 'Classe Concorso', 'Tipo Attività', 'Modalità', 'Sede', 'Note']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    esami = Esame.query.all()
    row_counter = 2  # Inizia dalla seconda riga dopo l'header
    
    for esame in esami:
        # Ottieni il nome della classe di concorso
        classe_concorso_nome = dict(CLASSI_CONCORSO).get(esame.classe_concorso, esame.classe_concorso)
        
        # Aggiungi data principale dell'esame
        if esame.data_inizio:
            ws.cell(row=row_counter, column=1, value=esame.data_inizio.strftime('%d/%m/%Y'))
            ws.cell(row=row_counter, column=2, value=esame.data_inizio.strftime('%H:%M'))
            ws.cell(row=row_counter, column=3, value=esame.data_fine.strftime('%H:%M') if esame.data_fine else "")
            ws.cell(row=row_counter, column=4, value=classe_concorso_nome)
            ws.cell(row=row_counter, column=5, value="Esame Principale")
            ws.cell(row=row_counter, column=6, value=esame.modalita)
            ws.cell(row=row_counter, column=7, value=esame.sede or '')
            ws.cell(row=row_counter, column=8, value=esame.note or '')
            row_counter += 1
        
        # Aggiungi le date del calendario esami (date aggiuntive)
        for data in esame.calendario_esami:
            attivita_nome = dict(TIPI_ATTIVITA).get(data.attivita, 'Altro')
            
            ws.cell(row=row_counter, column=1, value=data.data_inizio.strftime('%d/%m/%Y'))
            ws.cell(row=row_counter, column=2, value=data.data_inizio.strftime('%H:%M'))
            ws.cell(row=row_counter, column=3, value=data.data_fine.strftime('%H:%M'))
            ws.cell(row=row_counter, column=4, value=classe_concorso_nome)
            ws.cell(row=row_counter, column=5, value=attivita_nome)
            ws.cell(row=row_counter, column=6, value=data.modalita)
            ws.cell(row=row_counter, column=7, value=data.sede or '')
            ws.cell(row=row_counter, column=8, value="")  # Non ci sono note per le date aggiuntive
            
            row_counter += 1
    
    # Applica formattazione per migliorare la leggibilità
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Aggiungi filtri all'header
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{row_counter-1}"
    
    # Salva il file temporaneo
    with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xlsx') as f:
        wb.save(f.name)
        temp_path = f.name
    
    return send_file(temp_path, as_attachment=True, download_name='date_esami_pef.xlsx', 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/scarica_excel_commissioni')
def scarica_excel_commissioni():
    """Genera un file Excel con tutti i membri delle commissioni, un record per ogni classe di concorso"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commissioni Esami PEF"
    
    # Headers
    headers = ['Classe Concorso', 'Nome', 'Cognome', 'Tipologia', 'Ruolo', 'Profilo', 'Email', 'Telefono', 'Note']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Ottieni tutti gli esami con le loro commissioni
    esami = Esame.query.all()
    row_counter = 2  # Inizia dalla seconda riga dopo l'header
    
    for esame in esami:
        # Ottieni il nome della classe di concorso
        classe_concorso_nome = dict(CLASSI_CONCORSO).get(esame.classe_concorso, esame.classe_concorso)
        
        # Aggiungi ogni membro della commissione
        for commissione in esame.commissioni:
            ws.cell(row=row_counter, column=1, value=classe_concorso_nome)
            ws.cell(row=row_counter, column=2, value=commissione.nome_membro)
            ws.cell(row=row_counter, column=3, value=commissione.cognome_membro)
            ws.cell(row=row_counter, column=4, value='Titolare' if commissione.tipologia == 'T' else 'Supplente')
            ws.cell(row=row_counter, column=5, value=dict(RUOLI_COMMISSIONE).get(commissione.ruolo, commissione.ruolo))
            ws.cell(row=row_counter, column=6, value=dict(PROFILI_COMMISSIONE).get(commissione.profilo, commissione.profilo))
            ws.cell(row=row_counter, column=7, value=commissione.email or '')
            ws.cell(row=row_counter, column=8, value=commissione.telefono or '')
            ws.cell(row=row_counter, column=9, value=commissione.note or '')
            
            row_counter += 1
    
    # Applica formattazione per migliorare la leggibilità
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Aggiungi filtri all'header
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{row_counter-1}"
    
    # Salva il file temporaneo
    with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xlsx') as f:
        wb.save(f.name)
        temp_path = f.name
    
    return send_file(temp_path, as_attachment=True, download_name='commissioni_esami_pef.xlsx', 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Routes per la gestione dei membri delle commissioni
@app.route('/membri_commissioni')
def membri_commissioni():
    """Lista di tutti i membri delle commissioni registrati nel sistema"""
    search = request.args.get('search', '')
    profilo_filter = request.args.get('profilo', '')
    
    # Query base
    query = MembroCommissione.query
    
    # Applicazione filtri
    if search:
        query = query.filter(
            db.or_(
                MembroCommissione.nome.ilike(f'%{search}%'),
                MembroCommissione.cognome.ilike(f'%{search}%'),
                MembroCommissione.email.ilike(f'%{search}%')
            )
        )
    
    if profilo_filter:
        query = query.filter(MembroCommissione.profilo == profilo_filter)
    
    # Ordinamento per cognome e nome
    membri = query.order_by(MembroCommissione.cognome, MembroCommissione.nome).all()
    
    return render_template(
        'membri_commissioni.html', 
        membri=membri,
        profili_commissione=PROFILI_COMMISSIONE,
        search=search,
        profilo_filter=profilo_filter
    )

@app.route('/nuovo_membro_commissione', methods=['GET', 'POST'])
def nuovo_membro_commissione():
    """Aggiunta di un nuovo membro delle commissioni"""
    form = MembroCommissioneForm()
    if form.validate_on_submit():
        membro = MembroCommissione(
            nome=form.nome.data,
            cognome=form.cognome.data,
            profilo=form.profilo.data,
            email=form.email.data,
            telefono=form.telefono.data,
            note=form.note.data
        )
        db.session.add(membro)
        db.session.commit()
        flash('Membro della commissione aggiunto con successo!', 'success')
        return redirect(url_for('membri_commissioni'))
    
    return render_template('form_membro_commissione.html', form=form, title='Nuovo Membro Commissione')

@app.route('/membro_commissione/<int:id>')
def dettaglio_membro_commissione(id):
    """Pagina dettaglio di un membro delle commissioni"""
    membro = MembroCommissione.query.get_or_404(id)
    
    # Calcola statistiche per il membro
    esami = membro.get_esami()
    ore_totali = membro.calcola_ore_totali()
    
    # Raggruppa esami per classe di concorso
    esami_per_classe = {}
    for esame in esami:
        if esame.classe_concorso not in esami_per_classe:
            esami_per_classe[esame.classe_concorso] = []
        esami_per_classe[esame.classe_concorso].append(esame)
    
    return render_template(
        'dettaglio_membro_commissione.html',
        membro=membro,
        esami=esami,
        esami_per_classe=esami_per_classe,
        ore_totali=ore_totali,
        n_commissioni=len(membro.commissioni),
        n_esami=len(esami)
    )

@app.route('/membro_commissione/<int:id>/modifica', methods=['GET', 'POST'])
def modifica_membro_commissione(id):
    """Modifica di un membro delle commissioni"""
    membro = MembroCommissione.query.get_or_404(id)
    form = MembroCommissioneForm(obj=membro)
    
    if form.validate_on_submit():
        # Salva i dati precedenti per controllare cosa è cambiato
        vecchio_nome = membro.nome
        vecchio_cognome = membro.cognome
        vecchio_profilo = membro.profilo
        vecchia_email = membro.email
        vecchio_telefono = membro.telefono
        vecchie_note = membro.note
        
        # Aggiorna il membro
        form.populate_obj(membro)
        
        # Aggiorna anche tutte le commissioni associate se i dati sono cambiati
        commissioni = Commissione.query.filter_by(membro_id=membro.id).all()
        for commissione in commissioni:
            if vecchio_nome != membro.nome:
                commissione.nome_membro = membro.nome
            if vecchio_cognome != membro.cognome:
                commissione.cognome_membro = membro.cognome
            if vecchio_profilo != membro.profilo:
                commissione.profilo = membro.profilo
            if vecchia_email != membro.email:
                commissione.email = membro.email
            if vecchio_telefono != membro.telefono:
                commissione.telefono = membro.telefono
            if vecchie_note != membro.note:
                commissione.note = membro.note
        
        db.session.commit()
        flash('Membro della commissione aggiornato con successo!', 'success')
        return redirect(url_for('dettaglio_membro_commissione', id=membro.id))
    
    return render_template('form_membro_commissione.html', form=form, title='Modifica Membro Commissione')

@app.route('/membro_commissione/<int:id>/elimina', methods=['POST'])
def elimina_membro_commissione(id):
    """Eliminazione di un membro delle commissioni"""
    membro = MembroCommissione.query.get_or_404(id)
    nome_completo = f"{membro.nome} {membro.cognome}"
    
    # Verifica se il membro è associato a commissioni
    if membro.commissioni:
        flash(f'Impossibile eliminare: {nome_completo} è associato a {len(membro.commissioni)} commissioni', 'danger')
        return redirect(url_for('dettaglio_membro_commissione', id=membro.id))
    
    db.session.delete(membro)
    db.session.commit()
    flash(f'Membro della commissione {nome_completo} eliminato con successo!', 'success')
    
    return redirect(url_for('membri_commissioni'))

@app.route('/membro_commissione/<int:id>/aggiungi_commissioni', methods=['GET', 'POST'])
def aggiungi_membro_a_commissioni(id):
    """Aggiunta di un membro a multiple commissioni esistenti"""
    membro = MembroCommissione.query.get_or_404(id)
    
    # Ottenere gli ID degli esami in cui il membro è già presente
    esami_esistenti = [c.esame_id for c in membro.commissioni]
    
    # Esami disponibili (esclusi quelli in cui è già presente)
    esami_disponibili = Esame.query.filter(~Esame.id.in_(esami_esistenti)).all()
    
    if request.method == 'POST':
        esami_selezionati = request.form.getlist('esami')
        tipologia = request.form.get('tipologia', 'T')
        ruolo = request.form.get('ruolo')  # Nuovo ruolo specifico per questa commissione
        
        for esame_id in esami_selezionati:
            # Crea una nuova associazione commissione
            commissione = Commissione(
                esame_id=esame_id,
                membro_id=membro.id,
                nome_membro=membro.nome,
                cognome_membro=membro.cognome,
                tipologia=tipologia,
                ruolo=ruolo,  # Usa il ruolo specifico per questa commissione
                profilo=membro.profilo,
                email=membro.email,
                telefono=membro.telefono,
                note=membro.note
            )
            db.session.add(commissione)
        
        db.session.commit()
        flash(f'Membro aggiunto a {len(esami_selezionati)} commissioni!', 'success')
        return redirect(url_for('dettaglio_membro_commissione', id=membro.id))
    
    return render_template(
        'aggiungi_membro_commissioni.html',
        membro=membro,
        esami=esami_disponibili,
        tipologie=TIPOLOGIE_COMMISSIONE,
        ruoli=RUOLI_COMMISSIONE  # Aggiungi i ruoli disponibili
    )

# API Routes
@app.route('/api/classi-concorso')
def api_classi_concorso():
    """API per ottenere la lista delle classi di concorso"""
    return jsonify(dict(CLASSI_CONCORSO))

@app.route('/api/percorsi-pef')
def api_percorsi_pef():
    """API per ottenere la lista dei percorsi PEF"""
    return jsonify(dict(PERCORSI_PEF))

@app.route('/api/esami')
def api_esami():
    """API per ottenere la lista degli esami in formato JSON"""
    esami = Esame.query.all()
    esami_json = []
    for esame in esami:
        esame_data = {
            'id': esame.id,
            'classe_concorso': esame.classe_concorso,
            'percorsi_pef': esame.get_percorsi_list(),
            'data_inizio': esame.data_inizio.isoformat() if esame.data_inizio else None,
            'data_fine': esame.data_fine.isoformat() if esame.data_fine else None,
            'modalita': esame.modalita,
            'sede': esame.sede,
            'note': esame.note,
            'commissioni': [
                {
                    'id': c.id,
                    'nome_membro': c.nome_membro,
                    'cognome_membro': c.cognome_membro,
                    'ruolo': c.ruolo,
                    'profilo': c.profilo,
                    'email': c.email,
                    'telefono': c.telefono,
                    'note': c.note
                } for c in esame.commissioni
            ],
            'calendario_esami': [
                {
                    'data_inizio': d.data_inizio.isoformat(),
                    'data_fine': d.data_fine.isoformat(),
                    'modalita': d.modalita,
                    'sede': d.sede,
                    'attivita': d.attivita if hasattr(d, 'attivita') else 'altro'
                } for d in esame.calendario_esami
            ]
        }
        esami_json.append(esame_data)
    return jsonify(esami_json)

@app.route('/calendario')
def calendario():
    """Visualizza tutte le date degli esami in formato calendario"""
    # Ottieni tutti gli esami con le loro date principali e del calendario
    esami = Esame.query.all()
    
    # Prepara gli eventi per il calendario
    eventi = []
    
    for esame in esami:
        # Aggiungi data principale dell'esame
        if esame.data_inizio:
            classe_concorso_nome = dict(CLASSI_CONCORSO).get(esame.classe_concorso, esame.classe_concorso)
            percorsi_str = ', '.join([dict(PERCORSI_PEF).get(p, p) for p in esame.get_percorsi_list()])
            
            # Evento principale
            eventi.append({
                'id': f'esame_{esame.id}',
                'title': f"{classe_concorso_nome}",
                'description': f"{percorsi_str} - {esame.note or ''}",
                'start': esame.data_inizio.isoformat(),
                'end': esame.data_fine.isoformat() if esame.data_fine else None,
                'url': url_for('dettaglio_esame', id=esame.id),
                'color': '#3788d8',  # Blu per eventi principali
                'textColor': '#ffffff',
                'allDay': False
            })
        
        # Aggiungi date del calendario esami
        for data in esame.calendario_esami:
            attivita_nome = dict(TIPI_ATTIVITA).get(data.attivita, 'Altro')
            eventi.append({
                'id': f'calendario_{data.id}',
                'title': f"{classe_concorso_nome} - {attivita_nome}",
                'description': f"{percorsi_str} - {data.sede or ''}",
                'start': data.data_inizio.isoformat(),
                'end': data.data_fine.isoformat(),
                'url': url_for('dettaglio_esame', id=esame.id),
                'color': '#28a745',  # Verde per date aggiuntive
                'textColor': '#ffffff',
                'allDay': False
            })
    
    return render_template('calendario.html', eventi=eventi)

# Aggiungi questa funzione per verificare e aggiornare il database
def check_and_update_db():
    """Verifica se il database ha la struttura corretta e la aggiorna se necessario"""
    try:
        # Esegui la migrazione direttamente qui
        import sqlite3
        db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
        
        # Se è un percorso relativo, gestiscilo correttamente
        if not os.path.isabs(db_path):
            db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), db_path)
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Controlla se le colonne esistono già
        cursor.execute("PRAGMA table_info(esame)")
        existing_columns = [row[1] for row in cursor.fetchall()]
        
        columns_added = False
        
        if 'data_inizio' not in existing_columns:
            print("Aggiunta colonna data_inizio alla tabella esame...")
            cursor.execute("ALTER TABLE esame ADD COLUMN data_inizio TIMESTAMP")
            columns_added = True
        
        if 'data_fine' not in existing_columns:
            print("Aggiunta colonna data_fine alla tabella esame...")
            cursor.execute("ALTER TABLE esame ADD COLUMN data_fine TIMESTAMP")
            columns_added = True
        
        # Commit delle modifiche
        conn.commit()
        conn.close()
        
        if columns_added:
            print("Migrazione database completata con successo!")
        
        return True
    except Exception as migration_error:
        print(f"Errore durante la migrazione: {migration_error}")
        print("Continuando con funzionalità limitate...")
        return False

if __name__ == '__main__':
    # Esegui la migrazione del database prima di tutto il resto
    migration_success = check_and_update_db()
    
    # Poi crea le tabelle mancanti
    with app.app_context():
        db.create_all()
    
    app.run(debug=True)
